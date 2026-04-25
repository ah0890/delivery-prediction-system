from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from datetime import datetime, timedelta
import os
import logging
import re
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ================= CONFIGURATION =================
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Tracking Data")
os.makedirs(OUTPUT_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

DATE_PATTERN = re.compile(r'^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}')


def parse_tracking_format(tracking_num):
    """Extracts prefix, numeric part, suffix, and numeric length."""
    match = re.match(r"^([A-Za-z]+)(\d+)([A-Za-z]+)$", tracking_num)
    if not match:
        raise ValueError(f"Invalid format: '{tracking_num}'")
    return match.group(1), int(match.group(2)), match.group(3), len(match.group(2))


def generate_tracking_range(start_num, end_num):
    """Generates list of tracking numbers."""
    prefix_s, num_s, suffix_s, len_s = parse_tracking_format(start_num)
    prefix_e, num_e, suffix_e, len_e = parse_tracking_format(end_num)

    if prefix_s != prefix_e or suffix_s != suffix_e:
        raise ValueError("Start and end must have SAME prefix and suffix")
    if num_s > num_e:
        raise ValueError("Start number cannot be greater than end number")

    tracking_list = []
    for i in range(num_s, num_e + 1):
        padded_num = str(i).zfill(len_s)
        tracking_list.append(f"{prefix_s}{padded_num}{suffix_s}")
    return tracking_list


def parse_datetime(date_str):
    """Parse datetime from logistics record."""
    try:
        # Format: 2023-12-19 17:57(GMT+2)
        match = re.match(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2})', date_str)
        if match:
            return datetime.strptime(match.group(1), '%Y-%m-%d %H:%M')
    except:
        pass
    return None


def calculate_delivery_metrics(logistics_pairs):
    """Calculate delivery time metrics."""
    if not logistics_pairs:
        return None, None, None, "No Data"
    
    timestamps = []
    for ts, desc in logistics_pairs:
        dt = parse_datetime(ts)
        if dt:
            timestamps.append((dt, desc))
    
    if not timestamps:
        return None, None, None, "No Valid Dates"
    
    # Sort by datetime
    timestamps.sort(key=lambda x: x[0])
    
    order_confirmed = None
    client_received = None
    total_days = None
    status = "In Transit"
    
    # Find order confirmed and client received
    for dt, desc in timestamps:
        desc_lower = desc.lower()
        if 'order confirmed' in desc_lower or 'order placed' in desc_lower:
            order_confirmed = dt
        if 'client received' in desc_lower or 'delivered' in desc_lower or 'received' in desc_lower:
            client_received = dt
            status = "Delivered"
    
    # Calculate delivery time
    if order_confirmed and client_received:
        delta = client_received - order_confirmed
        total_days = delta.total_seconds() / 86400  # Convert to days
    elif order_confirmed:
        delta = datetime.now() - order_confirmed
        total_days = delta.total_seconds() / 86400
    
    return order_confirmed, client_received, total_days, status


def fetch_tracking_data(tracking_num):
    """Fetch tracking data and return as dictionary."""
    driver = None
    try:
        # ================= DRIVER SETUP =================
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-infobars")
        options.add_argument("--disable-extensions")

        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 30)

        # ================= OPEN & INTERACT =================
        driver.get("https://buffaloex.co.za/track.html")

        dropdown = wait.until(EC.presence_of_element_located((By.TAG_NAME, "select")))
        Select(dropdown).select_by_visible_text("Tracking No.")

        tracking_input = wait.until(EC.presence_of_element_located((By.ID, "order-no")))
        tracking_input.clear()
        tracking_input.send_keys(tracking_num)

        search_btn = wait.until(EC.element_to_be_clickable((By.ID, "submit-btn3")))
        driver.execute_script("arguments[0].click();", search_btn)

        wait.until(EC.visibility_of_element_located((By.ID, "search-result")))

        # ================= SCRAPE DATA =================
        result_content = driver.find_element(By.ID, "result-content")
        all_text = result_content.text
        lines = [line.strip() for line in all_text.split('\n') if line.strip()]

        receive_city = ""
        sign_status = "Not signed"
        logistics_pairs = []

        i = 0
        while i < len(lines):
            line = lines[i]

            if 'Receivecity' in line or 'Receive city' in line:
                receive_city = line.split(':')[-1].strip() if ':' in line else line
                i += 1; continue
            if 'Sign in picture' in line:
                sign_status = lines[i+1].strip() if i+1 < len(lines) else "Not signed"
                i += 2; continue
            if 'Logistics records' in line.lower():
                i += 1; continue

            if DATE_PATTERN.match(line):
                timestamp = line
                description = ""
                if i + 1 < len(lines):
                    next_line = lines[i+1]
                    if not DATE_PATTERN.match(next_line) and not any(h in next_line.lower() for h in ['receivecity', 'sign', 'logistics']):
                        description = next_line
                        i += 2
                    else:
                        i += 1
                else:
                    i += 1
                logistics_pairs.append((timestamp, description))
            else:
                i += 1

        # ================= CALCULATE METRICS =================
        order_date, delivery_date, total_days, delivery_status = calculate_delivery_metrics(logistics_pairs)
        
        # Determine if delayed (assuming > 7 days is delayed)
        is_delayed = "Yes" if total_days and total_days > 7 else "No" if total_days else "N/A"
        
        return {
            'tracking_number': tracking_num,
            'receive_city': receive_city,
            'sign_status': sign_status,
            'order_confirmed': order_date.strftime('%Y-%m-%d %H:%M') if order_date else "N/A",
            'delivered_date': delivery_date.strftime('%Y-%m-%d %H:%M') if delivery_date else "N/A",
            'total_days': round(total_days, 2) if total_days else "N/A",
            'delivery_status': delivery_status,
            'is_delayed': is_delayed,
            'logistics_count': len(logistics_pairs),
            'last_update': logistics_pairs[0][0] if logistics_pairs else "N/A",
            'error': None
        }

    except TimeoutException:
        return {'tracking_number': tracking_num, 'error': 'Timeout'}
    except Exception as e:
        return {'tracking_number': tracking_num, 'error': str(e)}
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass


def create_excel_report(data_list, output_file):
    """Create formatted Excel report with analysis."""
    
    # Create DataFrame
    df = pd.DataFrame(data_list)
    
    # Reorder columns
    columns = [
        'tracking_number', 'receive_city', 'delivery_status', 'is_delayed',
        'order_confirmed', 'delivered_date', 'total_days', 'sign_status',
        'logistics_count', 'last_update', 'error'
    ]
    df = df[columns]
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Tracking Data', index=False)
        
        # Create summary sheet
        summary_df = pd.DataFrame({
            'Metric': ['Total Packages', 'Delivered', 'In Transit', 'Delayed', 'On Time', 'Avg Delivery Days'],
            'Value': [
                len(df),
                len(df[df['delivery_status'] == 'Delivered']),
                len(df[df['delivery_status'] == 'In Transit']),
                len(df[df['is_delayed'] == 'Yes']),
                len(df[df['is_delayed'] == 'No']),
                df[df['total_days'] != 'N/A']['total_days'].astype(float).mean()
            ]
        })
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Create delayed packages sheet
        delayed_df = df[df['is_delayed'] == 'Yes'][columns]
        delayed_df.to_excel(writer, sheet_name='Delayed Packages', index=False)
        
        # Format sheets
        workbook = writer.book
        
        # Format main data sheet
        worksheet = writer.sheets['Tracking Data']
        
        # Add formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        delayed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        
        # Style headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Highlight delayed packages
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            if row[3].value == 'Yes':  # is_delayed column
                for cell in row:
                    cell.fill = delayed_fill
    
    logging.info(f"✅ Excel report saved: {output_file}")


# ================= MAIN EXECUTION =================
if __name__ == "__main__":
    print("📦 BuffaloEx Batch Tracker with Analysis")
    print("="*50)
    
    try:
        start_input = input("🔹 Enter START tracking number: ").strip()
        end_input = input("🔹 Enter END tracking number:   ").strip()

        if not start_input or not end_input:
            print("❌ Both start and end numbers are required.")
            exit(1)

        print("\n🔄 Generating tracking range...")
        tracking_list = generate_tracking_range(start_input, end_input)
        print(f"✅ Found {len(tracking_list)} tracking numbers to process.")
        print(f"📂 Output folder: {OUTPUT_DIR}\n")

        all_data = []
        success_count = 0
        fail_count = 0

        for i, num in enumerate(tracking_list, 1):
            print(f"[{i}/{len(tracking_list)}] Processing: {num}")
            data = fetch_tracking_data(num)
            all_data.append(data)
            
            if data.get('error') is None:
                success_count += 1
                status = data.get('delivery_status', 'Unknown')
                days = data.get('total_days', 'N/A')
                delayed = data.get('is_delayed', 'N/A')
                print(f"   ✅ Status: {status} | Days: {days} | Delayed: {delayed}")
            else:
                fail_count += 1
                print(f"   ❌ Error: {data.get('error')}")
            
            # Polite delay
            if i < len(tracking_list):
                time.sleep(2)

        # Save to Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(OUTPUT_DIR, f"Tracking_Analysis_{timestamp}.xlsx")
        create_excel_report(all_data, excel_file)

        # Also save CSV for compatibility
        csv_file = os.path.join(OUTPUT_DIR, f"Tracking_Analysis_{timestamp}.csv")
        pd.DataFrame(all_data).to_csv(csv_file, index=False)
        
        print("\n" + "="*50)
        print(f"🎉 BATCH COMPLETE!")
        print(f"✅ Successful: {success_count}")
        print(f"❌ Failed: {fail_count}")
        print(f"📊 Excel Report: {excel_file}")
        print(f"📄 CSV File: {csv_file}")
        print("="*50)

    except KeyboardInterrupt:
        print("\n⚠️ Process interrupted by user.")
    except ValueError as ve:
        print(f"\n❌ Format Error: {ve}")
    except Exception as e:
        print(f"\n❌ Unexpected Error: {e}")
        import traceback
        traceback.print_exc()